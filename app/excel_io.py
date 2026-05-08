# excel_io.py
"""Excel 导入 / 导出。"""
import io
from datetime import datetime, date
from openpyxl import load_workbook, Workbook


def parse_excel(file_bytes: bytes, expected_fields: list[str]) -> list[dict]:
    """
    解析上传的 xlsx，返回行数据列表。

    - 第一行为表头
    - 表头需包含 expected_fields 中的列（多余列忽略，缺失列以空值填充）
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        headers.append(str(cell).strip() if cell is not None else "")

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue  # 跳过空行
        record = {}
        for h, v in zip(headers, raw):
            if h not in expected_fields:
                continue
            record[h] = _normalize_value(v)
        # 缺失字段补空
        for f in expected_fields:
            record.setdefault(f, "")
        rows.append(record)
    return rows


def _normalize_value(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return v


def build_excel(field_names: list[str], rows: list[dict]) -> bytes:
    """把数据导出成 xlsx。"""
    wb = Workbook()
    ws = wb.active
    ws.append(field_names)
    for r in rows:
        ws.append([r.get(f, "") for f in field_names])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
